import openpyxl

# Orange color codes commonly used in Excel (may need to adjust if your orange is different)
ORANGE_RGBS = ["FFFF9900", "FFFFA500", "FFFF8000"]

def is_orange(cell):
    font = cell.font
    if font and font.color is not None:
        color = getattr(font.color, 'rgb', None)
        if color and color.upper() in ORANGE_RGBS:
            return True
    fill = cell.fill
    if fill and fill.fgColor is not None:
        color = getattr(fill.fgColor, 'rgb', None)
        if color and color.upper() in ORANGE_RGBS:
            return True
    return False

def extract_orange_theory_tasks(xlsx_path, output_path):
    wb = openpyxl.load_workbook(xlsx_path)
    if 'A+ 12.220' in wb.sheetnames:
        ws = wb['A+ 12.220']
    else:
        print("Sheet 'A+ 12.220' not found!")
        return
    orange_theory_tasks = []
    for row in ws.iter_rows():
        # Find the cell in the Practical/Theory column
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "Theory" in cell.value:
                if is_orange(cell):
                    # Get the corresponding Lab Title
                    lab_title = row[0].value if row[0].value else "(No Title)"
                    orange_theory_tasks.append(lab_title)
    orange_theory_tasks = [t for t in orange_theory_tasks if t and t.strip()]
    with open(output_path, 'w') as f:
        for task in orange_theory_tasks:
            f.write(str(task) + '\n')
    print(f"Extracted {len(orange_theory_tasks)} orange Theory tasks to {output_path}")

if __name__ == "__main__":
    extract_orange_theory_tasks("A+ 12.220.xlsx", "orange_theory_tasks.txt")
