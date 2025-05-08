import openpyxl

def find_theory_lines(xlsx_path, sheet_name, output_path):
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found!")
        return
    ws = wb[sheet_name]
    results = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "Theory" in cell.value:
                lab_title = row[0].value if row[0].value else "(No Title)"
                results.append(f"{lab_title} | {cell.value}")
    with open(output_path, 'w') as f:
        for line in results:
            f.write(line + '\n')
    print(f"Found {len(results)} lines with 'Theory' in {sheet_name}. Results written to {output_path}")

if __name__ == "__main__":
    find_theory_lines("A+ 12.220.xlsx", "A+ 12.220", "theory_lines.txt")
