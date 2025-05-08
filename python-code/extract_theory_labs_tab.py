import openpyxl

def extract_theory_labs_tab(xlsx_path, sheet_name, output_path):
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found!")
        return
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            rows.append(row)
    with open(output_path, 'w') as f:
        for row in rows:
            line = '\t'.join([str(cell) if cell is not None else '' for cell in row])
            f.write(line + '\n')
    print(f"Extracted {len(rows)} rows from '{sheet_name}' to {output_path}")

if __name__ == "__main__":
    extract_theory_labs_tab("Documents/A+ 12.220.xlsx", "Theory Labs", "A+ 220 Theory Labs.md")
