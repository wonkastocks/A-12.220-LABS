import openpyxl
import re
from pathlib import Path

# Load objectives from the extracted objectives.txt
with open('objectives.txt') as f:
    objectives = f.read().split('\n')

# Heuristic: Use the first objective block that matches keywords from the tab name or tasks
OBJECTIVE_BLOCKS = []
current_block = []
for line in objectives:
    if re.match(r'\d+\.\d+', line):
        if current_block:
            OBJECTIVE_BLOCKS.append(current_block)
        current_block = [line]
    else:
        if current_block:
            current_block.append(line)
if current_block:
    OBJECTIVE_BLOCKS.append(current_block)

def find_objective(tab, tasks):
    keywords = re.findall(r'\w+', tab.lower())
    for task in tasks:
        keywords += re.findall(r'\w+', str(task).lower())
    for block in OBJECTIVE_BLOCKS:
        block_text = ' '.join(block).lower()
        if any(word in block_text for word in keywords):
            return block
    return OBJECTIVE_BLOCKS[0] if OBJECTIVE_BLOCKS else []

wb = openpyxl.load_workbook('A+ 12.220-Rev1 (1).xlsx')
output_dir = Path('generated_lab_pages')
output_dir.mkdir(exist_ok=True)
index_links = []

for tab in wb.sheetnames:
    ws = wb[tab]
    tasks = [row[0].value for row in ws.iter_rows(min_row=1, max_col=1) if row[0].value]
    if not tasks:
        continue
    # Find correlated objective
    objective_block = find_objective(tab, tasks)
    # Write HTML page
    fname = re.sub(r'[^\w\-]+', '_', tab.strip()) + '.html'
    index_links.append(f'<li><a href="{fname}">{tab}</a></li>')
    with open(output_dir / fname, 'w') as f:
        f.write(f"""<!DOCTYPE html>
<html lang='en'>
<head>
  <meta charset='UTF-8'>
  <title>Lab: {tab}</title>
  <style>body {{ font-family: Arial, sans-serif; margin: 2em; }} h1 {{ color: #2c3e50; }} ul {{ margin-bottom: 2em; }} .objectives {{ background: #f9f9f9; border-left: 4px solid #2c3e50; padding: 1em; margin-bottom: 1em; }}</style>
</head>
<body>
  <h1>Lab: {tab}</h1>
  <h2>Tasks to be Covered</h2>
  <ul>
    {''.join(f'<li>{task}</li>' for task in tasks)}
  </ul>
  <h2>Correlated 12.220 Objectives</h2>
  <div class='objectives'>
    {'<br/>'.join(objective_block)}
  </div>
</body>
</html>""")

# Write index.html
with open(output_dir / 'index.html', 'w') as f:
    f.write(f"""<!DOCTYPE html>
<html lang='en'>
<head><meta charset='UTF-8'><title>Lab Index</title></head>
<body>
  <h1>A+ 12.220 Labs Index</h1>
  <ul>
    {''.join(index_links)}
  </ul>
</body>
</html>""")
