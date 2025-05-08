import openpyxl
import re
from pathlib import Path

# Load objectives from the extracted objectives.txt
with open('objectives.txt') as f:
    objectives = f.read().split('\n')

# Parse objectives into blocks with number and name
OBJECTIVE_BLOCKS = []
current_block = []
for line in objectives:
    if re.match(r'(\d+\.\d+)', line):
        if current_block:
            OBJECTIVE_BLOCKS.append(current_block)
        current_block = [line]
    else:
        if current_block:
            current_block.append(line)
if current_block:
    OBJECTIVE_BLOCKS.append(current_block)

def find_objective(task):
    task_lc = str(task).lower()
    for block in OBJECTIVE_BLOCKS:
        block_text = ' '.join(block).lower()
        if any(word in block_text for word in re.findall(r'\w+', task_lc)):
            return block[0], block[1] if len(block) > 1 else ''
    # fallback to first objective
    return OBJECTIVE_BLOCKS[0][0], OBJECTIVE_BLOCKS[0][1] if len(OBJECTIVE_BLOCKS[0]) > 1 else ''

wb = openpyxl.load_workbook('A+ 12.220-Rev1 (1).xlsx')
output_dir = Path('.')

for tab in wb.sheetnames:
    ws = wb[tab]
    tasks = [row[0].value for row in ws.iter_rows(min_row=1, max_col=1) if row[0].value]
    if not tasks:
        continue
    fname = f"Lab_{re.sub(r'[^\w\-]+', '_', tab.strip())}.md"
    with open(output_dir / fname, 'w') as f:
        f.write(f"# {tab}\n\n## Task List\n")
        for task in tasks:
            f.write(f"- {task}\n")
        f.write("\n## Tasks to be Covered and Correlated Objectives\n\n")
        for task in tasks:
            obj_num, obj_name = find_objective(task)
            f.write(f"- **{task}**  \n  - {obj_num}: {obj_name}\n")
        f.write("\n---\n")
